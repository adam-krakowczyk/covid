from os.path import isfile
from openpyxl import Workbook, load_workbook
from pathlib import Path
from os import walk

SOURCE_EXCEL_1 = 'test_kwarantanna1.xlsx'
SOURCE_EXCEL_2 = 'test_wynikidodatnie1.xlsx'
DESTINATION_EXCEL ='test_wynik_pasuje.xlsx'


def sum_excels_from_catalog(catalog):
    data ={}
    global_row_id = 1
    try:
        _, _, filenames = next(walk(catalog))
        print(filenames)
        for file in filenames:
            xlsx_file = Path('sprawdzeni', file)
            wb_source = load_workbook(xlsx_file)
            sheet_source = wb_source.active
            for row_id in range(2, sheet_source.max_row+1):
                data_row=[]
                for col_id in range(1, sheet_source.max_column+1):
                    cell_obj = sheet_source.cell(row=row_id, column=col_id)
                    data_row.append(cell_obj.value)
                data[global_row_id]=data_row
                global_row_id+=1

        #print(data)
        wb_destination = Workbook()
        sheet_destination = wb_destination.active
        sheet_destination.title = 'scalony excel'
        for key,value in data.items():
            sheet_destination.append(value)
            #print(value)
        wb_destination.save(filename='zeszyt_all.xlsx')

    except:
        print('brak plików w katalogu')



def xls_compare(file1,file2):
        '''
        funkcja porównuje adresy z zpliku file1 z plikiem file2, czyści pole adres z smieci
        :param file1:
        :param file2:
        :return: none
        '''

        print ('[checking files exist]')
        if not isfile(file1):
            print('[-] File doesnt exist: ', file1)
            return
        if not isfile(file2):
            print('[-] File doesnt exist: ', file2)
            return
        print('[+] Files exist')

        xlsx_file = Path('', file1)
        wb_obj = load_workbook(xlsx_file)

        # Read the active sheet:
        sheet = wb_obj.active

        adres_1=[]

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:      # nagłówek
                pass
            else:
                adres_1.append(str(row[4]).upper().replace('UL. ','').replace('AL. ','').replace('PL. ','').replace(', KATOWICE','').replace('ALEJA ','').strip())
        print(adres_1)
        print('-'*20)
        print('załadowano z pliku : ',file1, ' adresów: ', len(adres_1))
        print('-'*20)

        xlsx_file2 = Path('', file2)
        wb_obj2 = load_workbook(xlsx_file2)

        # Read the active sheet:
        sheet2 = wb_obj2.active
        adres_2=[]

        for i, row in enumerate(sheet2.iter_rows(values_only=True)):
            if i == 0:
                pass
            else:
                if len(row[6])>0 and str(row[6]).isalnum():
                        adres_2.append(str(row[4]).upper().replace('UL. ','').replace('AL. ','').replace('\\','').replace('_','').lstrip()
                               + ' '
                               + row[5]
                               + '/'
                               + row[6])
                elif len(row[6])==0:
                        adres_2.append(str(row[4]).upper().replace('UL. ','').replace('AL. ','').replace('\\','').replace('_','').lstrip()
                               + ' '
                               + row[5])

        print(adres_2)
        print('-'*20)
        print('załadowano z pliku : ',file2, ' adresów: ', len(adres_2))
        print('-'*20)
        adres_to_call =[]
        for adres in adres_1:
            if adres in adres_2:
                adres_to_call.append(adres)
        print('Spójne adresy: ', len(adres_to_call))
        print(adres_to_call)

        #zapis do xlsx
        wb = Workbook()

        ws1 = wb.active
        ws1.title = 'Pasujące adresy'
        ws1['A1']='Adres'
        i=1
        for row in adres_to_call:
            i+=1
            ws1.cell(column=1,row=i,value=str(row).capitalize())
        wb.save(filename=DESTINATION_EXCEL)


#xls_compare(SOURCE_EXCEL_1,SOURCE_EXCEL_2)
sum_excels_from_catalog('sprawdzeni')
